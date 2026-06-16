import pandas as pd
from fastapi.testclient import TestClient

from backend.app.schemas.dictionaries import DictionaryType
from backend.app.schemas.files import FileKind
from backend.app.main import app
from backend.app.repositories.job_repository import job_repository
from backend.app.services import file_classifier
from backend.app.services.column_normalizer import normalize_dataframe
from backend.app.services.file_classifier import detect_dictionary, inspect_file


REPORT_COLUMNS = ["Type", "Last - Title", "First", "NPI", "CBcode", "Practice", "DOS", "SIN"]


def test_usap_providers_detected_by_columns() -> None:
    df = normalize_dataframe(pd.DataFrame(columns=["NAME", "NPI_NUMBER", "ProvMnemonic", "BA_MNEMONIC"]))
    detection = detect_dictionary(df)
    assert detection.detected_type == DictionaryType.USAP_PROVIDERS
    assert detection.missing_columns == []


def test_referring_providers_detected_by_columns() -> None:
    df = normalize_dataframe(pd.DataFrame(columns=["NAME", "NUMBER", "NPI_NUMBER", "Lastname", "Firstname"]))
    detection = detect_dictionary(df)
    assert detection.detected_type == DictionaryType.REFERRING_PROVIDERS


def test_report_workbook_without_correction_signals_is_report(tmp_path) -> None:
    path = tmp_path / "report.xlsx"
    pd.DataFrame([["MD", "DOE", "JANE", "1234567890", "", "P1", "6/1/2026", "SIN1"]], columns=REPORT_COLUMNS).to_excel(path, index=False)

    inspection = inspect_file(path, "file1", "anything.xlsx")

    assert inspection.kind == FileKind.CB_FAILED_REPORT


def test_temporary_spreadsheet_lock_file_is_ignored(tmp_path) -> None:
    path = tmp_path / "~$FINELLI FL FAILED TO UPLOAD REPORT.xlsx"
    path.write_bytes(b"temporary lock")

    inspection = inspect_file(path, "file1", path.name)

    assert inspection.kind == FileKind.IGNORE
    assert inspection.row_count == 0
    assert inspection.warnings == ["Temporary or system file ignored."]


def test_report_columns_with_chg_to_signal_is_corrections(tmp_path) -> None:
    path = tmp_path / "corrections.xlsx"
    pd.DataFrame([["MD", "ABRAHAM", "JEBY", "CHG TO WING", "MD9019", "P1", "6/1/2026", "SIN1"]], columns=REPORT_COLUMNS).to_excel(path, index=False)

    inspection = inspect_file(path, "file1", "anything.xlsx")

    assert inspection.kind == FileKind.CORRECTIONS


def test_comments_and_source_values_make_corrections(tmp_path) -> None:
    path = tmp_path / "source.xlsx"
    df = pd.DataFrame(
        [["MD", "DOE", "JANE", "", "", "P1", "6/1/2026", "SIN1", "Change in the ticket", "Dictionary"]],
        columns=[*REPORT_COLUMNS, "Comments", "Source"],
    )
    df.to_excel(path, index=False)

    inspection = inspect_file(path, "file1", "anything.xlsx")

    assert inspection.kind == FileKind.CORRECTIONS


def test_correction_formatting_makes_corrections(tmp_path) -> None:
    path = tmp_path / "formatted.xlsx"
    pd.DataFrame([["MD", "DOE", "JANE", "1234567890", "", "P1", "6/1/2026", "SIN1"]], columns=REPORT_COLUMNS).to_excel(path, index=False)
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet["D2"].font = Font(color="00FF0000")
    workbook.save(path)

    inspection = inspect_file(path, "file1", "anything.xlsx")

    assert inspection.kind == FileKind.CORRECTIONS


def test_multi_sheet_formatting_inspection_reuses_workbook(tmp_path, monkeypatch) -> None:
    path = tmp_path / "formatted_multi_sheet.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([["MD", "DOE", "JANE", "1234567890", "", "P1", "6/1/2026", "SIN1"]], columns=REPORT_COLUMNS).to_excel(
            writer,
            sheet_name="Original",
            index=False,
        )
        pd.DataFrame([["MD", "SMITH", "ALICE", "1234567890", "", "P1", "6/1/2026", "SIN2"]], columns=REPORT_COLUMNS).to_excel(
            writer,
            sheet_name="Corrections",
            index=False,
        )

    from openpyxl import load_workbook as real_load_workbook
    from openpyxl.styles import Font

    workbook = real_load_workbook(path)
    worksheet = workbook["Corrections"]
    worksheet["D2"].font = Font(color="00FF0000")
    workbook.save(path)
    workbook.close()

    calls = []

    def counted_load_workbook(*args, **kwargs):
        calls.append(args[0])
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(file_classifier, "load_workbook", counted_load_workbook)

    inspection = inspect_file(path, "file1", "anything.xlsx")

    assert inspection.kind == FileKind.CORRECTIONS
    assert calls == [path]


def test_create_job_applies_file_type_override() -> None:
    client = TestClient(app)
    response = client.post(
        "/api/uploads/inspect",
        files=[("files", ("ambiguous.xlsx", b"not really excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert response.status_code == 200
    payload = response.json()
    upload_id = payload["upload_id"]
    file_id = payload["files"][0]["file_id"]

    create_response = client.post("/api/jobs", json={"upload_id": upload_id, "file_overrides": {file_id: "IGNORE"}})

    assert create_response.status_code == 200
    upload = job_repository.get_upload(upload_id)
    assert upload.files[0].inspection.kind == FileKind.IGNORE
