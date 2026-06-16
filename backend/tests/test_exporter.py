from io import BytesIO

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.main import app
from backend.app.repositories.job_repository import job_repository
from backend.app.schemas.ai import AIAction, AIInterpretation, AIReasonCode
from backend.app.schemas.results import RowDetail, ValidationResult, ValidationStatus
from backend.app.services.excel_exporter import rows_to_workbook

EXPECTED_CLEAN_COLUMNS = [
    "SIN",
    "Row",
    "Type",
    "Action",
    "Apply",
    "Status",
    "Current Provider",
    "Current NPI",
    "Current CBCode",
    "Recommended Provider",
    "Recommended NPI",
    "Recommended CBCode",
    "Comments",
    "Source",
    "Reason",
]

EXPECTED_FINAL_COLUMNS = [
    "Type",
    "Last - Title",
    "First",
    "NPI",
    "CBcode",
    "Comments",
    "Source",
    "Practice",
    "DOS",
    "SIN",
]

TECHNICAL_COLUMNS = [
    "Bot_Accion",
    "AI_Action",
    "AI_Reason_Code",
    "Validation_Details",
    "Cell_Color_NPI",
    "Cell_Color_CBCode",
]


def _row(
    *,
    row_id: str = "r1",
    sin: str = "SIN1",
    region: str = "MARYLAND",
    final_action: str = "COMPLETE_INFO",
    quick_action: str = "Complete fields",
    apply_this: str = "YES",
    source: str = "Dictionary",
    role: str = "Provider",
    current_last: str = "DOE",
    current_first: str = "JANE",
    recommended_last: str = "SMITH",
    recommended_first: str = "ALICE",
    manual_reason: str = "",
    correction_summary: str = "Dictionary validated CBCode CB1.",
) -> RowDetail:
    interp = AIInterpretation(
        action=AIAction.COMPLETE_INFO,
        reason_code=AIReasonCode.DIRECT_NPI,
        target_provider_name=None,
        target_npi="1234567890",
        target_cbcode=None,
        requires_add_to_ge=False,
        is_pending_usap=False,
        confidence=1,
        needs_manual_review=False,
        explanation="ok",
    )
    validation = ValidationResult(status=ValidationStatus.NPI_FOUND, details="ok", matches=[], npi_registry_name=None, needs_manual_review=False)
    return RowDetail(
        row_id=row_id,
        sheet_name="Sheet1",
        SIN=sin,
        Region=region,
        Row_Index=8,
        sanitized_original={"npi": "1234567890", "practice": "MCD Shared", "dos": "2026-05-21"},
        Bot_Accion="COMPLETE_INFO",
        Bot_Suggestion="ok",
        Bot_Details="ok",
        AI_Action="COMPLETE_INFO",
        AI_Reason_Code="DIRECT_NPI",
        AI_Confidence=1,
        Needs_Manual_Review=False,
        Validation_Status="NPI_FOUND",
        Validation_Details="ok",
        Dictionary_Match_Type=None,
        Matched_Dictionary=None,
        Matched_NPI=None,
        Matched_CBCode=None,
        Matched_Provider_Name=None,
        Deactivation_Status=None,
        AI_Explanation="ok",
        Final_Action=final_action,
        Final_Recommendation="ok",
        Quick_Action=quick_action,
        Apply_This=apply_this,
        Current_Type=role,
        Recommended_Type=role,
        Current_Last_Title=current_last,
        Current_First=current_first,
        Current_NPI="1234567890",
        Current_CBCode="",
        Recommended_Last_Title=recommended_last,
        Recommended_First=recommended_first,
        Recommended_NPI="1987654321",
        Recommended_CBCode="CB1",
        Recommended_Comments="Change in the ticket",
        Recommended_Source=source,
        Correction_Summary=correction_summary,
        Manual_Reason=manual_reason,
        Cell_Color_Last_Title="red",
        Cell_Color_First="red",
        Cell_Color_NPI="green",
        Cell_Color_CBCode="green",
        Cell_Color_Comments="red",
        Cell_Color_Source="green",
        deterministic_interpretation=interp,
        ai_interpretation=interp,
        validation=validation,
    )


def test_summary_export_returns_workbook_bytes() -> None:
    row = _row()
    assert rows_to_workbook([row], kind="summary").startswith(b"PK")


def test_apply_ready_export_filters_yes_rows() -> None:
    data = rows_to_workbook(
        [_row(row_id="ready", sin="READY", apply_this="YES"), _row(row_id="hold", sin="HOLD", apply_this="NO")],
        kind="apply_ready",
    )
    df = pd.read_excel(BytesIO(data))
    assert df["SIN"].tolist() == ["READY"]
    assert list(df.columns) == EXPECTED_CLEAN_COLUMNS


def test_usap_export_filters_awaiting_and_clears_source() -> None:
    data = rows_to_workbook(
        [
            _row(row_id="apply", sin="APPLY", final_action="COMPLETE_INFO", apply_this="YES", source="Dictionary"),
            _row(row_id="usap", sin="USAP", final_action="AWAITING_USAP", apply_this="NO", source="Dictionary"),
        ],
        kind="usap",
    )
    df = pd.read_excel(BytesIO(data)).fillna("")
    assert df["SIN"].tolist() == ["USAP"]
    assert df.loc[0, "Source"] == ""


def test_numbers_ready_export_is_clean_and_grouped_by_region() -> None:
    data = rows_to_workbook(
        [
            _row(row_id="md", sin="MD1", region="MARYLAND", apply_this="YES"),
            _row(row_id="tx", sin="TX1", region="TEXAS", apply_this="NO", final_action="MANUAL_REVIEW", quick_action="Manual review"),
        ],
        kind="numbers_ready",
    )
    workbook = load_workbook(BytesIO(data))
    assert workbook.sheetnames[:4] == ["Summary", "Apply Ready", "MARYLAND", "TEXAS"]

    df = pd.read_excel(BytesIO(data), sheet_name="MARYLAND")
    assert list(df.columns) == EXPECTED_FINAL_COLUMNS
    assert all(column not in df.columns for column in TECHNICAL_COLUMNS)
    assert df.loc[0, "Type"] == "Provider"
    assert df.loc[0, "Last - Title"] == "DOE"
    assert df.loc[0, "First"] == "JANE"
    assert str(df.loc[0, "NPI"]) == "1987654321"
    assert df.loc[0, "CBcode"] == "CB1"
    assert df.loc[0, "Practice"] == "MCD Shared"
    assert str(df.loc[0, "DOS"]) == "2026-05-21"
    assert df.loc[0, "SIN"] == "MD1"

    apply_ready = pd.read_excel(BytesIO(data), sheet_name="Apply Ready")
    assert list(apply_ready.columns) == EXPECTED_FINAL_COLUMNS
    assert apply_ready["SIN"].tolist() == ["MD1"]

    summary = pd.read_excel(BytesIO(data), sheet_name="Summary")
    assert {"Total rows", "Ready to Apply", "Change Ticket", "Complete Fields"}.issubset(set(summary["Metric"]))


def test_numbers_ready_styles_change_ticket_with_rich_text_without_color_columns() -> None:
    data = rows_to_workbook(
        [
            _row(
                row_id="md",
                apply_this="YES",
                final_action="CHANGE_TICKET",
                quick_action="Change ticket",
                current_last="TANG",
                current_first="AN THIEN",
                recommended_last="TANG",
                recommended_first="ANDREW",
            )
        ],
        kind="numbers_ready",
    )
    workbook = load_workbook(BytesIO(data), rich_text=True)
    worksheet = workbook["MARYLAND"]
    headers = [cell.value for cell in worksheet[1]]
    assert "Cell_Color_NPI" not in headers
    assert "Cell_Color_CBCode" not in headers
    assert "Current Provider" not in headers
    assert "Recommended Provider" not in headers

    columns = {header: index + 1 for index, header in enumerate(headers)}
    last_cell = worksheet.cell(row=2, column=columns["Last - Title"])
    first_cell = worksheet.cell(row=2, column=columns["First"])
    npi_cell = worksheet.cell(row=2, column=columns["NPI"])
    assert str(last_cell.value) == "TANG TANG"
    assert str(first_cell.value) == "AN THIEN ANDREW"
    assert last_cell.value[0].font.strike is True
    assert last_cell.value[2].font.color.rgb == "FFFF0000"
    assert first_cell.value[0].font.strike is True
    assert first_cell.value[2].font.color.rgb == "FFFF0000"
    assert npi_cell.value[0].font.strike is True
    assert npi_cell.value[2].font.color.rgb == "FFFF0000"
    assert worksheet.cell(row=2, column=columns["Comments"]).value[0].font.color.rgb == "FFFF0000"


def test_numbers_ready_complete_info_does_not_use_rich_strikethrough() -> None:
    data = rows_to_workbook([_row(row_id="md", apply_this="YES")], kind="numbers_ready")
    workbook = load_workbook(BytesIO(data), rich_text=True)
    worksheet = workbook["MARYLAND"]
    headers = [cell.value for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    last_cell = worksheet.cell(row=2, column=columns["Last - Title"])
    assert last_cell.value == "DOE"
    assert getattr(last_cell.value, "font", None) is None


def test_numbers_ready_complete_info_surgeon_colors_added_name_green() -> None:
    data = rows_to_workbook(
        [
            _row(
                row_id="surgeon-complete",
                role="Surgeon",
                current_last="HARDY",
                current_first="ELVIN",
                recommended_last="HARDY",
                recommended_first="ELVIN KENDELL",
            )
        ],
        kind="numbers_ready",
    )
    workbook = load_workbook(BytesIO(data), rich_text=True)
    worksheet = workbook["MARYLAND"]
    headers = [cell.value for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    first_cell = worksheet.cell(row=2, column=columns["First"])
    assert str(first_cell.value) == "ELVIN KENDELL"
    assert first_cell.value[0] == "ELVIN"
    assert first_cell.value[1].text == " KENDELL"
    assert first_cell.value[1].font.color.rgb == "FF008000"
    assert str(worksheet.cell(row=2, column=columns["CBcode"]).value) == "CB1"
    assert worksheet.cell(row=2, column=columns["CBcode"]).value[0].font.color.rgb == "FF008000"


def test_numbers_ready_complete_info_provider_keeps_system_name() -> None:
    data = rows_to_workbook(
        [
            _row(
                row_id="provider-complete",
                role="Provider",
                current_last="Guidry",
                current_first="Xenequia Monique",
                recommended_last="GUIDRY",
                recommended_first="XENEQUIA MONIQUE",
            )
        ],
        kind="numbers_ready",
    )
    df = pd.read_excel(BytesIO(data), sheet_name="MARYLAND")
    assert df.loc[0, "Last - Title"] == "Guidry"
    assert df.loc[0, "First"] == "Xenequia Monique"


def test_full_export_download_does_not_delete_job_files(tmp_path) -> None:
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    full_export_path = job_dir / "processed_full.xlsx"
    full_export_path.write_bytes(b"raw workbook bytes")
    job = job_repository.create_job("export-no-delete", "upload", job_dir)
    job.full_export_path = str(full_export_path)
    client = TestClient(app)

    try:
        response = client.get("/api/export/export-no-delete", params={"kind": "full"})

        assert response.status_code == 200
        assert full_export_path.exists()
        assert job_dir.exists()
        assert job_repository.get_job("export-no-delete") is not None
    finally:
        job_repository.jobs.pop("export-no-delete", None)
