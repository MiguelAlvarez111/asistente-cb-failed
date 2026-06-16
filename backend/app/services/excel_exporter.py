from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.schemas.results import RowDetail


ANALYST_COLUMNS = [
    "SIN",
    "Region",
    "Row_Index",
    "Final_Action",
    "Quick_Action",
    "Apply_This",
    "Current_Type",
    "Recommended_Type",
    "Current_Last_Title",
    "Current_First",
    "Current_NPI",
    "Current_CBCode",
    "Recommended_Last_Title",
    "Recommended_First",
    "Recommended_NPI",
    "Recommended_CBCode",
    "Recommended_Comments",
    "Recommended_Source",
    "Correction_Summary",
    "Analyst_Next_Step",
    "Needs_Manual_Review",
    "Manual_Reason",
    "Validation_Status",
    "Matched_Dictionary",
    "Matched_Provider_Name",
    "Matched_NPI",
    "Matched_CBCode",
    "AI_Confidence",
    "Cell_Color_Last_Title",
    "Cell_Color_First",
    "Cell_Color_NPI",
    "Cell_Color_CBCode",
    "Cell_Color_Comments",
    "Cell_Color_Source",
]

LEGACY_COLUMNS = [
    "Bot_Accion",
    "Bot_Suggestion",
    "Bot_Details",
    "AI_Action",
    "AI_Reason_Code",
    "Validation_Details",
    "Dictionary_Match_Type",
    "Deactivation_Status",
    "AI_Explanation",
    "Final_Recommendation",
]

OUTPUT_COLUMNS = ANALYST_COLUMNS + LEGACY_COLUMNS

FILL_COLORS = {
    "red": "FFFFC7CE",
    "green": "FFC6EFCE",
    "yellow": "FFFFEB9C",
    "gray": "FFE7E6E6",
    "neutral": "FFE7E6E6",
}

COLOR_TO_VALUE_COLUMN = {
    "Cell_Color_Last_Title": "Recommended_Last_Title",
    "Cell_Color_First": "Recommended_First",
    "Cell_Color_NPI": "Recommended_NPI",
    "Cell_Color_CBCode": "Recommended_CBCode",
    "Cell_Color_Comments": "Recommended_Comments",
    "Cell_Color_Source": "Recommended_Source",
}

CLEAN_COLUMNS = [
    "SIN",
    "Row",
    "Type",
    "Action",
    "Apply",
    "Status",
    "Current Provider",
    "Current NPI",
    "Current CBCode",
    "Recommended Provider",
    "Recommended NPI",
    "Recommended CBCode",
    "Comments",
    "Source",
    "Reason",
]

FINAL_DELIVERY_COLUMNS = [
    "Type",
    "Last - Title",
    "First",
    "NPI",
    "CBcode",
    "Comments",
    "Source",
    "Practice",
    "DOS",
    "SIN",
]

CLEAN_COLUMN_WIDTHS = {
    "SIN": 22,
    "Row": 8,
    "Type": 12,
    "Action": 18,
    "Apply": 18,
    "Status": 14,
    "Current Provider": 28,
    "Current NPI": 16,
    "Current CBCode": 16,
    "Recommended Provider": 28,
    "Recommended NPI": 16,
    "Recommended CBCode": 18,
    "Comments": 34,
    "Source": 16,
    "Reason": 44,
}

FINAL_COLUMN_WIDTHS = {
    "Type": 12,
    "Last - Title": 28,
    "First": 38,
    "NPI": 26,
    "CBcode": 22,
    "Comments": 28,
    "Source": 18,
    "Practice": 22,
    "DOS": 14,
    "SIN": 34,
}

ACTION_LABELS = {
    "CHANGE_TICKET": "Change Ticket",
    "COMPLETE_INFO": "Complete Fields",
    "AWAITING_USAP": "Awaiting USAP",
    "MANUAL_REVIEW": "Manual Review",
    "REMOVE_FROM_TICKET": "Remove from Ticket",
    "MALFORMED_ROW": "Invalid Row",
    "ADD_TO_GE": "Awaiting USAP",
    "NO_ACTION": "No Action",
}

SUMMARY_ACTIONS = [
    ("Change Ticket", "CHANGE_TICKET"),
    ("Complete Fields", "COMPLETE_INFO"),
    ("Awaiting USAP", "AWAITING_USAP"),
    ("Manual Review", "MANUAL_REVIEW"),
    ("Remove from Ticket", "REMOVE_FROM_TICKET"),
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF21453E")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
SUBTLE_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFF7FAF8")
THIN_BORDER = Border(bottom=Side(style="thin", color="FFE0E5E1"))
MONO_FONT = Font(name="Courier New")
MONO_COLUMNS = {"SIN", "Current NPI", "Current CBCode", "Recommended NPI", "Recommended CBCode"}
CENTER_COLUMNS = {"Row", "Type", "Action", "Apply", "Status"}
WRAP_COLUMNS = {"Current Provider", "Recommended Provider", "Comments", "Reason"}
FINAL_WRAP_COLUMNS = {"Last - Title", "First", "Comments", "Practice", "SIN"}
FINAL_MONO_COLUMNS = {"NPI", "CBcode", "SIN"}
FINAL_CENTER_COLUMNS = {"Type", "DOS"}


def _export_row(row: RowDetail) -> dict[str, object]:
    data = row.model_dump(
        mode="json",
        exclude={"sanitized_original", "deterministic_interpretation", "ai_interpretation", "validation", "correction_instruction"},
    )
    export_row: dict[str, object] = {"row_id": data.get("row_id", ""), "sheet_name": data.get("sheet_name", "")}
    export_row.update({column: data.get(column, "") for column in OUTPUT_COLUMNS})
    return export_row


def _provider_display(last_title: str | None, first: str | None) -> str:
    last = str(last_title or "").strip()
    first_name = str(first or "").strip()
    if last and first_name:
        return f"{last}, {first_name}"
    return last or first_name


def _action_label(row: RowDetail) -> str:
    return row.Quick_Action or ACTION_LABELS.get(row.Final_Action, row.Final_Action)


def _apply_label(row: RowDetail) -> str:
    return "Ready to Apply" if row.Apply_This == "YES" else "Do Not Apply Yet"


def _work_status_label(row: RowDetail) -> str:
    return row.Work_Status.value if hasattr(row.Work_Status, "value") else str(row.Work_Status or "")


def _reason(row: RowDetail) -> str:
    return row.Manual_Reason or row.Correction_Summary or ""


def _is_surgeon(row: RowDetail) -> bool:
    return str(row.Current_Type or row.Recommended_Type or "").strip().lower() == "surgeon"


def _is_provider(row: RowDetail) -> bool:
    return str(row.Current_Type or row.Recommended_Type or "").strip().lower() == "provider"


def _original_value(row: RowDetail, key: str) -> str:
    original = row.sanitized_original or {}
    return str(original.get(key, "") or "").strip()


def _recommended_or_current(recommended: str | None, current: str | None) -> str:
    recommended_text = str(recommended or "").strip()
    current_text = str(current or "").strip()
    return recommended_text or current_text


def _combined_plain(current: str | None, recommended: str | None, *, include_current_when_same: bool = True) -> str:
    current_text = str(current or "").strip()
    recommended_text = str(recommended or "").strip()
    if current_text and recommended_text:
        if current_text == recommended_text and not include_current_when_same:
            return recommended_text
        return f"{current_text} {recommended_text}"
    return recommended_text or current_text


def _final_delivery_row(row: RowDetail) -> dict[str, object]:
    is_change_ticket = row.Final_Action == "CHANGE_TICKET"
    current_last = row.Current_Last_Title
    current_first = row.Current_First
    current_npi = row.Current_NPI
    current_cbcode = row.Current_CBCode
    recommended_last = row.Recommended_Last_Title
    recommended_first = row.Recommended_First
    recommended_npi = row.Recommended_NPI
    recommended_cbcode = row.Recommended_CBCode

    if is_change_ticket:
        last_title = _combined_plain(current_last, recommended_last)
        first = _combined_plain(current_first, recommended_first)
        npi = _combined_plain(current_npi, recommended_npi)
        cbcode = _combined_plain(current_cbcode, recommended_cbcode)
    elif _is_provider(row):
        last_title = _recommended_or_current(current_last, recommended_last)
        first = _recommended_or_current(current_first, recommended_first)
        npi = _recommended_or_current(recommended_npi, current_npi)
        cbcode = _recommended_or_current(recommended_cbcode, current_cbcode)
    else:
        last_title = _recommended_or_current(recommended_last, current_last)
        first = _recommended_or_current(recommended_first, current_first)
        npi = _recommended_or_current(recommended_npi, current_npi)
        cbcode = _recommended_or_current(recommended_cbcode, current_cbcode)

    return {
        "Type": row.Current_Type or row.Recommended_Type,
        "Last - Title": last_title,
        "First": first,
        "NPI": npi,
        "CBcode": cbcode,
        "Comments": row.Recommended_Comments,
        "Source": row.Recommended_Source,
        "Practice": _original_value(row, "practice"),
        "DOS": _original_value(row, "dos"),
        "SIN": row.SIN,
    }


def _rich_change_value(current: str | None, recommended: str | None, *, include_current_when_same: bool = True) -> CellRichText | str:
    current_text = str(current or "").strip()
    recommended_text = str(recommended or "").strip()
    if current_text and recommended_text:
        if current_text == recommended_text and not include_current_when_same:
            return recommended_text
        return CellRichText(
            [
                TextBlock(InlineFont(strike=True), current_text),
                " ",
                TextBlock(InlineFont(color="FFFF0000"), recommended_text),
            ]
        )
    if recommended_text:
        return CellRichText([TextBlock(InlineFont(color="FFFF0000"), recommended_text)])
    return current_text


def _rich_red_value(value: str | None) -> CellRichText | str:
    text = str(value or "").strip()
    if not text:
        return ""
    return CellRichText([TextBlock(InlineFont(color="FFFF0000"), text)])


def _rich_green_value(value: str | None) -> CellRichText | str:
    text = str(value or "").strip()
    if not text:
        return ""
    return CellRichText([TextBlock(InlineFont(color="FF008000"), text)])


def _rich_added_value(current: str | None, recommended: str | None) -> CellRichText | str:
    current_text = str(current or "").strip()
    recommended_text = str(recommended or "").strip()
    if not recommended_text:
        return current_text
    if not current_text or len(recommended_text) <= len(current_text):
        return recommended_text
    if not recommended_text.upper().startswith(current_text.upper()):
        return recommended_text
    base = recommended_text[: len(current_text)]
    added = recommended_text[len(current_text) :]
    if not added.strip():
        return recommended_text
    return CellRichText([base, TextBlock(InlineFont(color="FF008000"), added)])


def _clean_export_row(row: RowDetail, *, source_override: str | None = None) -> dict[str, object]:
    source = row.Recommended_Source if source_override is None else source_override
    return {
        "SIN": row.SIN,
        "Row": row.Row_Index,
        "Type": row.Current_Type or row.Recommended_Type,
        "Action": _action_label(row),
        "Apply": _apply_label(row),
        "Status": _work_status_label(row),
        "Current Provider": _provider_display(row.Current_Last_Title, row.Current_First),
        "Current NPI": row.Current_NPI,
        "Current CBCode": row.Current_CBCode,
        "Recommended Provider": _provider_display(row.Recommended_Last_Title, row.Recommended_First),
        "Recommended NPI": row.Recommended_NPI,
        "Recommended CBCode": row.Recommended_CBCode,
        "Comments": row.Recommended_Comments,
        "Source": source,
        "Reason": _reason(row),
    }


def _first_action_color(colors: list[str]) -> str:
    normalized = {color.lower() for color in colors if color}
    for color in ["red", "yellow", "green", "gray", "neutral"]:
        if color in normalized:
            return color
    return ""


def _clean_color_context(row: RowDetail, *, source_override: str | None = None) -> dict[str, str]:
    source_color = "gray" if source_override == "" else row.Cell_Color_Source
    return {
        "Recommended Provider": _first_action_color([row.Cell_Color_Last_Title, row.Cell_Color_First]),
        "Recommended NPI": row.Cell_Color_NPI,
        "Recommended CBCode": row.Cell_Color_CBCode,
        "Comments": row.Cell_Color_Comments,
        "Source": source_color,
    }


def _clean_frame(rows: list[RowDetail], *, source_override: str | None = None) -> tuple[pd.DataFrame, list[dict[str, str]]]:
    return (
        pd.DataFrame([_clean_export_row(row, source_override=source_override) for row in rows], columns=CLEAN_COLUMNS),
        [_clean_color_context(row, source_override=source_override) for row in rows],
    )


def _final_delivery_frame(rows: list[RowDetail]) -> pd.DataFrame:
    return pd.DataFrame([_final_delivery_row(row) for row in rows], columns=FINAL_DELIVERY_COLUMNS)


def _apply_color_styles(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    worksheet = writer.sheets[sheet_name]
    columns = {column: index + 1 for index, column in enumerate(df.columns)}
    for color_column, value_column in COLOR_TO_VALUE_COLUMN.items():
        if color_column not in columns or value_column not in columns:
            continue
        color_index = columns[color_column]
        value_index = columns[value_column]
        for row_number in range(2, len(df) + 2):
            color_name = str(worksheet.cell(row=row_number, column=color_index).value or "").lower()
            rgb = FILL_COLORS.get(color_name)
            if rgb:
                fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
                worksheet.cell(row=row_number, column=value_index).fill = fill
                worksheet.cell(row=row_number, column=color_index).fill = fill


def _style_clean_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, row_colors: list[dict[str, str]]) -> None:
    worksheet = writer.sheets[sheet_name]
    columns = {column: index + 1 for index, column in enumerate(df.columns)}
    last_column = get_column_letter(max(len(df.columns), 1))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{last_column}{max(len(df) + 1, 1)}"

    for index, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(index)].width = CLEAN_COLUMN_WIDTHS.get(column, 16)

    for row_number in range(2, len(df) + 2):
        if row_number % 2 == 0:
            for column_index in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_number, column=column_index).fill = SUBTLE_ROW_FILL
        worksheet.row_dimensions[row_number].height = 24
        for column_name, column_index in columns.items():
            cell = worksheet.cell(row=row_number, column=column_index)
            horizontal = "center" if column_name in CENTER_COLUMNS else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=column_name in WRAP_COLUMNS)
            cell.border = THIN_BORDER
            if column_name in MONO_COLUMNS:
                cell.font = MONO_FONT

        color_context = row_colors[row_number - 2]
        for column_name, color_name in color_context.items():
            if column_name not in columns:
                continue
            rgb = FILL_COLORS.get(str(color_name or "").lower())
            if rgb:
                worksheet.cell(row=row_number, column=columns[column_name]).fill = PatternFill(
                    fill_type="solid",
                    fgColor=rgb,
                )


def _style_final_delivery_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, rows: list[RowDetail]) -> None:
    worksheet = writer.sheets[sheet_name]
    columns = {column: index + 1 for index, column in enumerate(df.columns)}
    last_column = get_column_letter(max(len(df.columns), 1))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{last_column}{max(len(df) + 1, 1)}"

    for index, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(index)].width = FINAL_COLUMN_WIDTHS.get(column, 16)

    for row_number in range(2, len(df) + 2):
        if row_number % 2 == 0:
            for column_index in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_number, column=column_index).fill = SUBTLE_ROW_FILL
        worksheet.row_dimensions[row_number].height = 24
        for column_name, column_index in columns.items():
            cell = worksheet.cell(row=row_number, column=column_index)
            horizontal = "center" if column_name in FINAL_CENTER_COLUMNS else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=column_name in FINAL_WRAP_COLUMNS)
            cell.border = THIN_BORDER
            if column_name in FINAL_MONO_COLUMNS:
                cell.font = MONO_FONT
                cell.number_format = "@"

        source_row = rows[row_number - 2]
        if source_row.Final_Action == "CHANGE_TICKET":
            rich_values = {
                "Last - Title": _rich_change_value(source_row.Current_Last_Title, source_row.Recommended_Last_Title),
                "First": _rich_change_value(source_row.Current_First, source_row.Recommended_First),
                "NPI": _rich_change_value(source_row.Current_NPI, source_row.Recommended_NPI),
                "CBcode": _rich_change_value(source_row.Current_CBCode, source_row.Recommended_CBCode),
                "Comments": _rich_red_value(source_row.Recommended_Comments),
            }
            for column_name, rich_value in rich_values.items():
                if column_name not in columns:
                    continue
                cell = worksheet.cell(row=row_number, column=columns[column_name])
                cell.value = rich_value
        elif source_row.Final_Action == "COMPLETE_INFO":
            if _is_surgeon(source_row):
                name_values = {
                    "Last - Title": _rich_added_value(source_row.Current_Last_Title, source_row.Recommended_Last_Title),
                    "First": _rich_added_value(source_row.Current_First, source_row.Recommended_First),
                }
                for column_name, rich_value in name_values.items():
                    if column_name in columns:
                        worksheet.cell(row=row_number, column=columns[column_name]).value = rich_value
            if source_row.Cell_Color_NPI.lower() == "green" and "NPI" in columns:
                worksheet.cell(row=row_number, column=columns["NPI"]).value = _rich_green_value(_recommended_or_current(source_row.Recommended_NPI, source_row.Current_NPI))
            if source_row.Cell_Color_CBCode.lower() == "green" and "CBcode" in columns:
                worksheet.cell(row=row_number, column=columns["CBcode"]).value = _rich_green_value(_recommended_or_current(source_row.Recommended_CBCode, source_row.Current_CBCode))
        elif source_row.Final_Action == "AWAITING_USAP" and "CBcode" in columns:
            worksheet.cell(row=row_number, column=columns["CBcode"]).fill = PatternFill(fill_type="solid", fgColor=FILL_COLORS["yellow"])


def _write_clean_sheet(writer: pd.ExcelWriter, sheet_name: str, rows: list[RowDetail], *, source_override: str | None = None) -> None:
    df, row_colors = _clean_frame(rows, source_override=source_override)
    safe_sheet_name = sheet_name[:31]
    df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
    _style_clean_sheet(writer, safe_sheet_name, df, row_colors)


def _write_final_delivery_sheet(writer: pd.ExcelWriter, sheet_name: str, rows: list[RowDetail]) -> None:
    df = _final_delivery_frame(rows)
    safe_sheet_name = sheet_name[:31]
    df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
    _style_final_delivery_sheet(writer, safe_sheet_name, df, rows)


def _summary_rows(rows: list[RowDetail]) -> list[dict[str, object]]:
    work_status_counts: dict[str, int] = {}
    for row in rows:
        status = _work_status_label(row)
        work_status_counts[status] = work_status_counts.get(status, 0) + 1

    summary = [
        {"Metric": "Total rows", "Count": len(rows)},
        {"Metric": "Ready to Apply", "Count": sum(1 for row in rows if row.Apply_This == "YES")},
    ]
    summary.extend(
        {"Metric": label, "Count": sum(1 for row in rows if row.Final_Action == action)}
        for label, action in SUMMARY_ACTIONS
    )
    summary.extend(
        {"Metric": label, "Count": work_status_counts.get(label, 0)}
        for label in ["Applied", "Copied", "Pending", "Skipped"]
    )
    return summary


def _write_summary_sheet(writer: pd.ExcelWriter, rows: list[RowDetail]) -> None:
    df = pd.DataFrame(_summary_rows(rows), columns=["Metric", "Count"])
    df.to_excel(writer, index=False, sheet_name="Summary")
    worksheet = writer.sheets["Summary"]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:B{max(len(df) + 1, 1)}"
    worksheet.column_dimensions["A"].width = 26
    worksheet.column_dimensions["B"].width = 12
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_number in range(2, len(df) + 2):
        worksheet.cell(row=row_number, column=1).border = THIN_BORDER
        worksheet.cell(row=row_number, column=2).border = THIN_BORDER
        worksheet.cell(row=row_number, column=2).alignment = Alignment(horizontal="center")
        if row_number % 2 == 0:
            worksheet.cell(row=row_number, column=1).fill = SUBTLE_ROW_FILL
            worksheet.cell(row=row_number, column=2).fill = SUBTLE_ROW_FILL


def write_processed_workbook(processed_sheets: dict[str, pd.DataFrame], output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in processed_sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])


def rows_to_workbook(rows: list[RowDetail], *, kind: str) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if kind == "summary":
            _write_summary_sheet(writer, rows)
        elif kind == "numbers_ready":
            _write_summary_sheet(writer, rows)
            _write_final_delivery_sheet(writer, "Apply Ready", [row for row in rows if row.Apply_This == "YES"])
            grouped: dict[str, list[RowDetail]] = {}
            for row in rows:
                region = str(row.Region or row.sheet_name or "Results")
                grouped.setdefault(region[:31], []).append(row)
            if not grouped:
                grouped["Results"] = []
            for sheet_name, group_rows in grouped.items():
                _write_final_delivery_sheet(writer, sheet_name, group_rows)
        elif kind == "full":
            df = pd.DataFrame([_export_row(row) for row in rows])
            df.to_excel(writer, index=False, sheet_name="full")
            _apply_color_styles(writer, "full", df)
        else:
            filtered_rows = rows
            source_override = None
            if kind == "manual_review":
                filtered_rows = [row for row in rows if row.Needs_Manual_Review]
            elif kind == "high_confidence":
                filtered_rows = [row for row in rows if row.AI_Confidence >= 0.9 and not row.Needs_Manual_Review]
            elif kind == "apply_ready":
                filtered_rows = [row for row in rows if row.Apply_This == "YES"]
            elif kind == "usap":
                filtered_rows = [row for row in rows if row.Final_Action == "AWAITING_USAP"]
                source_override = ""
            _write_clean_sheet(writer, kind[:31], filtered_rows, source_override=source_override)
    return output.getvalue()
