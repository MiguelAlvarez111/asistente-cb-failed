from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from backend.app.schemas.dictionaries import DictionaryDetection, DictionaryType
from backend.app.schemas.files import FileInspection, FileKind
from backend.app.services.column_normalizer import normalize_column_name, normalize_dataframe

USAP_PROVIDER_REQUIRED = {"npi_number", "prov_mnemonic", "ba_mnemonic"}
REFERRING_REQUIRED = {"npi_number", "number", "last_name", "first_name"}
REPORT_REQUIRED = {"type", "last_title", "first", "npi", "cbcode", "practice", "dos", "sin"}
SOURCE_SIGNALS = ["dictionary", "usap", "npi registry"]
FORMAT_COLUMNS = {"last_title", "first", "npi", "cbcode", "comments", "source"}


def detect_dictionary(df: pd.DataFrame) -> DictionaryDetection:
    columns = set(df.columns)
    warnings: list[str] = []
    if USAP_PROVIDER_REQUIRED.issubset(columns):
        missing: list[str] = []
        detected = DictionaryType.USAP_PROVIDERS
        confidence = 0.99
    elif REFERRING_REQUIRED.issubset(columns):
        missing = []
        detected = DictionaryType.REFERRING_PROVIDERS
        confidence = 0.98
    else:
        provider_missing = sorted(USAP_PROVIDER_REQUIRED - columns)
        referring_missing = sorted(REFERRING_REQUIRED - columns)
        missing = provider_missing if len(provider_missing) <= len(referring_missing) else referring_missing
        detected = DictionaryType.UNKNOWN
        confidence = 0.2
        warnings.append("Dictionary schema was not recognized from columns.")
    return DictionaryDetection(
        detected_type=detected,
        confidence=confidence,
        columns_found=sorted(columns),
        missing_columns=missing,
        row_count=len(df.index),
        warnings=warnings,
    )


def _read_file(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".txt":
        return pd.read_csv(path, sep="|", header=0, encoding="latin1", low_memory=False, dtype=str)
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.read_excel(path, sheet_name=0, dtype=str)
    return pd.DataFrame()


def _series_has_text(series: pd.Series, signals: list[str]) -> bool:
    for value in series.fillna("").astype(str):
        text = value.lower().strip()
        if text and any(signal in text for signal in signals):
            return True
    return False


def _has_non_empty_values(df: pd.DataFrame, columns: set[str]) -> bool:
    for column in columns & set(df.columns):
        if df[column].fillna("").astype(str).map(str.strip).astype(bool).any():
            return True
    return False


def _rgb_is_red_or_green(rgb: str) -> bool:
    value = rgb[-6:].upper()
    if len(value) != 6:
        return False
    try:
        red = int(value[0:2], 16)
        green = int(value[2:4], 16)
        blue = int(value[4:6], 16)
    except ValueError:
        return False
    return (red >= 150 and green <= 120 and blue <= 120) or (green >= 120 and red <= 140 and blue <= 140)


def _cell_has_correction_format(cell: Cell) -> bool:
    if cell.value in {None, ""}:
        return False
    if cell.font and cell.font.strike:
        return True
    color = cell.font.color if cell.font else None
    if color and color.type == "rgb" and color.rgb and _rgb_is_red_or_green(color.rgb):
        return True
    fill = cell.fill
    if fill and fill.fill_type and fill.fgColor:
        rgb = fill.fgColor.rgb
        indexed = fill.fgColor.indexed
        if rgb and str(rgb).upper() not in {"00000000", "FFFFFFFF", "00FFFFFF"}:
            return True
        if indexed not in {None, 64}:
            return True
    return False


def _sheet_has_correction_formatting(path: Path, sheet_name: str) -> bool:
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
        worksheet = workbook[sheet_name]
    except Exception:
        return False
    try:
        header = [normalize_column_name(cell.value) for cell in worksheet[1]]
        watched_indexes = [index + 1 for index, column in enumerate(header) if column in FORMAT_COLUMNS]
        if not watched_indexes:
            return False
        for row in worksheet.iter_rows(min_row=2):
            for column_index in watched_indexes:
                if column_index <= len(row) and _cell_has_correction_format(row[column_index - 1]):
                    return True
        return False
    finally:
        workbook.close()


def has_correction_signals(df: pd.DataFrame, *, path: Path | None = None, sheet_name: str | None = None) -> bool:
    columns = set(df.columns)
    if "npi" in columns and _series_has_text(df["npi"], ["chg to"]):
        return True
    if "cbcode" in columns and _series_has_text(df["cbcode"], ["chg to", "add to ge", "awaiting"]):
        return True
    if "comments" in columns and _series_has_text(
        df["comments"],
        ["awaiting", "change in the ticket", "remove from the ticket", "correct provider", "chg to", "add to ge"],
    ):
        return True
    if "source" in columns and _series_has_text(df["source"], SOURCE_SIGNALS):
        return True
    if _has_non_empty_values(df, {"comments", "source"}):
        return True
    if path and sheet_name and _sheet_has_correction_formatting(path, sheet_name):
        return True
    return False


def _inspect_excel(path: Path) -> FileInspection | None:
    workbook = pd.ExcelFile(path)
    report_columns: set[str] = set()
    correction_columns: set[str] = set()
    report_rows = 0
    correction_rows = 0
    report_sheets: list[str] = []
    correction_sheets: list[str] = []
    first_df: pd.DataFrame | None = None
    correction_signal_warnings: list[str] = []

    for sheet_name in workbook.sheet_names:
        df = normalize_dataframe(pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna(""))
        if first_df is None:
            first_df = df
        columns = set(df.columns)
        sheet_has_report_columns = REPORT_REQUIRED.issubset(columns)
        sheet_has_correction_signals = has_correction_signals(df, path=path, sheet_name=sheet_name)
        if sheet_has_correction_signals:
            correction_rows += len(df.index)
            correction_columns.update(columns)
            correction_sheets.append(sheet_name)
            if sheet_has_report_columns:
                report_rows += len(df.index)
                report_columns.update(columns)
                report_sheets.append(sheet_name)
            correction_signal_warnings.append(f"Correction signals detected in sheet: {sheet_name}")
        elif sheet_has_report_columns:
            report_rows += len(df.index)
            report_columns.update(columns)
            report_sheets.append(sheet_name)

    if correction_sheets:
        return FileInspection(
            file_id="",
            filename="",
            kind=FileKind.CORRECTIONS,
            row_count=correction_rows,
            column_count=len(correction_columns),
            columns_found=sorted(correction_columns),
            missing_columns=sorted({"sin"} - correction_columns),
            warnings=correction_signal_warnings,
            dictionary_detection=None,
        )

    if report_sheets:
        skipped = [sheet for sheet in workbook.sheet_names if sheet not in report_sheets]
        warnings = [f"Skipped non-report sheets: {', '.join(skipped)}"] if skipped else []
        return FileInspection(
            file_id="",
            filename="",
            kind=FileKind.CB_FAILED_REPORT,
            row_count=report_rows,
            column_count=len(report_columns),
            columns_found=sorted(report_columns),
            missing_columns=[],
            warnings=warnings,
            dictionary_detection=None,
        )

    if first_df is not None:
        return FileInspection(
            file_id="",
            filename="",
            kind=FileKind.UNKNOWN,
            row_count=len(first_df.index),
            column_count=len(first_df.columns),
            columns_found=sorted(first_df.columns),
            missing_columns=sorted(REPORT_REQUIRED - set(first_df.columns)),
            warnings=["File type was not recognized from workbook sheets."],
            dictionary_detection=None,
        )
    return None


def inspect_file(path: Path, file_id: str, filename: str) -> FileInspection:
    warnings: list[str] = []
    if filename == ".DS_Store" or filename.startswith("~$"):
        return FileInspection(
            file_id=file_id,
            filename=filename,
            kind=FileKind.IGNORE,
            row_count=0,
            column_count=0,
            columns_found=[],
            missing_columns=[],
            warnings=["Temporary or system file ignored."],
            dictionary_detection=None,
        )
    try:
        if path.suffix.lower() in {".xlsx", ".xls"}:
            inspection = _inspect_excel(path)
            if inspection is not None:
                return inspection.model_copy(update={"file_id": file_id, "filename": filename})
        df = normalize_dataframe(_read_file(path))
    except Exception as exc:
        return FileInspection(
            file_id=file_id,
            filename=filename,
            kind=FileKind.UNKNOWN,
            row_count=0,
            column_count=0,
            columns_found=[],
            missing_columns=[],
            warnings=[f"Unable to read file: {exc}"],
            dictionary_detection=None,
        )

    columns = set(df.columns)
    dictionary_detection = detect_dictionary(df) if path.suffix.lower() == ".txt" else None
    if dictionary_detection and dictionary_detection.detected_type != DictionaryType.UNKNOWN:
        kind = FileKind.DICTIONARY
        missing = dictionary_detection.missing_columns
    elif has_correction_signals(df):
        kind = FileKind.CORRECTIONS
        missing = sorted({"sin"} - columns)
    elif REPORT_REQUIRED.issubset(columns):
        kind = FileKind.CB_FAILED_REPORT
        missing = []
    else:
        kind = FileKind.UNKNOWN
        missing = sorted(REPORT_REQUIRED - columns)
        warnings.append("File type was not recognized from columns.")

    return FileInspection(
        file_id=file_id,
        filename=filename,
        kind=kind,
        row_count=len(df.index),
        column_count=len(df.columns),
        columns_found=sorted(columns),
        missing_columns=missing,
        warnings=warnings,
        dictionary_detection=dictionary_detection,
    )
